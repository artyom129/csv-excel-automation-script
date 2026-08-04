from pathlib import Path
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference


INPUT_FILE = Path("sample_orders.csv")
OUTPUT_DIR = Path("output")
OUTPUT_FILE = OUTPUT_DIR / "clean_sales_report.xlsx"


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def normalize_name(value: str) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().title()


def normalize_email(value: str) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def normalize_phone(value: str) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def load_data(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    df = pd.read_csv(path)
    required_columns = {
        "order_id", "date", "customer_name", "email",
        "phone", "product", "quantity", "price"
    }

    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"Missing columns: {', '.join(sorted(missing))}")

    return df


def clean_data(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    df = df.copy()

    df["customer_name"] = df["customer_name"].apply(normalize_name)
    df["email"] = df["email"].apply(normalize_email)
    df["phone"] = df["phone"].apply(normalize_phone)
    df["product"] = df["product"].astype(str).str.strip()

    df["quantity"] = pd.to_numeric(df["quantity"], errors="coerce")
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.drop_duplicates(subset=["order_id"], keep="first")

    df["error"] = ""
    df.loc[df["customer_name"].eq(""), "error"] += "Missing customer name; "
    df.loc[~df["email"].apply(lambda x: bool(EMAIL_RE.match(str(x)))), "error"] += "Invalid email; "
    df.loc[df["date"].isna(), "error"] += "Invalid date; "
    df.loc[df["quantity"].isna() | (df["quantity"] <= 0), "error"] += "Invalid quantity; "
    df.loc[df["price"].isna() | (df["price"] <= 0), "error"] += "Invalid price; "

    df["total"] = df["quantity"] * df["price"]

    invalid_df = df[df["error"].ne("")].copy()
    clean_df = df[df["error"].eq("")].copy()

    clean_df["date"] = clean_df["date"].dt.strftime("%Y-%m-%d")
    invalid_df["date"] = invalid_df["date"].dt.strftime("%Y-%m-%d")

    return clean_df, invalid_df


def build_dashboard(clean_df: pd.DataFrame, invalid_df: pd.DataFrame) -> pd.DataFrame:
    total_orders = len(clean_df)
    invalid_rows = len(invalid_df)
    total_revenue = clean_df["total"].sum() if total_orders else 0
    avg_order_value = clean_df["total"].mean() if total_orders else 0
    unique_customers = clean_df["email"].nunique() if total_orders else 0

    return pd.DataFrame({
        "Metric": [
            "Valid orders",
            "Invalid rows",
            "Unique customers",
            "Total revenue",
            "Average order value",
        ],
        "Value": [
            total_orders,
            invalid_rows,
            unique_customers,
            round(total_revenue, 2),
            round(avg_order_value, 2),
        ]
    })


def export_excel(clean_df: pd.DataFrame, invalid_df: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    dashboard_df = build_dashboard(clean_df, invalid_df)

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        dashboard_df.to_excel(writer, sheet_name="Dashboard", index=False)
        clean_df.to_excel(writer, sheet_name="Clean Data", index=False)
        invalid_df.to_excel(writer, sheet_name="Invalid Rows", index=False)

    format_workbook()


def format_workbook() -> None:
    wb = load_workbook(OUTPUT_FILE)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(size=16, bold=True, color="1F4E78")
    thin = Side(border_style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))

            ws.column_dimensions[column_letter].width = min(max_length + 3, 35)

    dashboard = wb["Dashboard"]
    dashboard.insert_rows(1)
    dashboard["A1"] = "CSV / Excel Automation Report"
    dashboard["A1"].font = title_font
    dashboard.merge_cells("A1:B1")
    dashboard["A1"].alignment = Alignment(horizontal="center")

    chart = BarChart()
    chart.title = "Report Metrics"
    chart.y_axis.title = "Value"
    chart.x_axis.title = "Metric"

    data = Reference(dashboard, min_col=2, min_row=3, max_row=7)
    cats = Reference(dashboard, min_col=1, min_row=3, max_row=7)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    dashboard.add_chart(chart, "D3")

    wb.save(OUTPUT_FILE)


def main() -> None:
    raw_df = load_data(INPUT_FILE)
    clean_df, invalid_df = clean_data(raw_df)
    export_excel(clean_df, invalid_df)

    print(f"Done. Report created: {OUTPUT_FILE}")
    print(f"Valid rows: {len(clean_df)}")
    print(f"Invalid rows: {len(invalid_df)}")


if __name__ == "__main__":
    main()
